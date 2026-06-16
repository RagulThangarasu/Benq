#!/usr/bin/env python3
"""Folder-wise CONTENT validation -> single Excel workbook.

Produces one workbook with:
  * Tab 1 "Consolidated" - all folders summary
  * Tabs 2..N based on folder name - per-folder validation details

Output: content_validation/folderwise_reports/content_validation_folderwise.xlsx
"""
import json
import re
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
OUT_DIR = HERE / "folderwise_reports"
SUMMARY_JSON = OUT_DIR / "summary.json"
OUT_XLSX = OUT_DIR / "content_validation_folderwise.xlsx"

import openpyxl  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HDR_FILL = PatternFill("solid", fgColor="34495E")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def _sheet_name(name, used):
    safe = re.sub(r"[\[\]:*?/\\]", "-", name).strip()[:31] or "sheet"
    base, i = safe, 1
    while safe.lower() in used:
        suffix = f"~{i}"
        safe = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(safe.lower())
    return safe


def _style_header(ws, row, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w


def _hyperlink_cell(ws, row, col, text, target):
    cell = ws.cell(row=row, column=col, value=text)
    cell.hyperlink = target
    cell.font = Font(color="0563C1", underline="single")
    cell.alignment = WRAP_TOP
    cell.border = BORDER
    return cell


def main():
    if not SUMMARY_JSON.exists():
        print(f"Missing summary JSON: {SUMMARY_JSON}")
        raise SystemExit(1)

    data = json.loads(SUMMARY_JSON.read_text())
    rows = data.get("results", [])
    if not rows:
        print(f"No results found in {SUMMARY_JSON}")
        raise SystemExit(1)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    cons = wb.active
    cons.title = "Consolidated"

    used_names = {"consolidated"}
    detail_tabs = []

    for idx, row in enumerate(rows, start=1):
        folder = row.get("folder", f"item_{idx}")
        sname = _sheet_name(folder, used_names)
        detail_tabs.append((sname, row))
        _write_detail_tab(wb, sname, row)

    _write_consolidated(cons, detail_tabs)
    wb.save(OUT_XLSX)
    print(f"Saved combined Excel workbook: {OUT_XLSX}")
    print(f"Sheets: 1 Consolidated + {len(detail_tabs)} folder tabs")


def _write_consolidated(ws, detail_tabs):
    ws.cell(row=1, column=1, value="Content Validation — Consolidated Product Report").font = TITLE_FONT
    total_products = len(detail_tabs)
    pass_count = sum(1 for _, row in detail_tabs if row.get("status") == "PASS")
    fail_count = sum(1 for _, row in detail_tabs if row.get("status") == "FAIL")
    toc_missing_total = sum(row.get("toc_missing", 0) for _, row in detail_tabs)
    content_fail_total = sum(row.get("content_fail", 0) for _, row in detail_tabs)

    summary_labels = [
        ("Total products", total_products),
        ("PASS", pass_count),
        ("FAIL", fail_count),
        ("Total TOC missing", toc_missing_total),
        ("Total content fail", content_fail_total),
    ]
    row_offset = 2
    for idx, (label, value) in enumerate(summary_labels, start=row_offset):
        ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=value)

    headers = [
        "#", "Product", "Status", "Prod PDF", "Stage PDF", "TOC match",
        "TOC missing", "TOC extra", "Content pass", "Content fail", "Report"
    ]
    widths = [4, 32, 12, 30, 30, 12, 12, 12, 14, 14, 38]
    header_row = row_offset + len(summary_labels) + 1
    _style_header(ws, header_row, headers, widths)

    for r, (sname, row) in enumerate(detail_tabs, start=header_row + 1):
        ws.cell(row=r, column=1, value=r - header_row)
        fcell = ws.cell(row=r, column=2)
        fcell.value = row.get("folder", "")
        fcell.hyperlink = f"#'{sname}'!A1"
        fcell.font = Font(color="0563C1", underline="single")
        v = row.get("status", "")
        ws.cell(row=r, column=3, value=v)
        ws.cell(row=r, column=4, value=row.get("prod_pdf", ""))
        ws.cell(row=r, column=5, value=row.get("stage_pdf", ""))
        ws.cell(row=r, column=6, value=row.get("toc_match", ""))
        ws.cell(row=r, column=7, value=row.get("toc_missing", ""))
        ws.cell(row=r, column=8, value=row.get("toc_extra", ""))
        ws.cell(row=r, column=9, value=row.get("content_pass", ""))
        ws.cell(row=r, column=10, value=row.get("content_fail", ""))
        report_path = row.get("report", "")
        if report_path:
            _hyperlink_cell(ws, r, 11, "Open PDF", Path(report_path).as_uri())
        else:
            ws.cell(row=r, column=11, value="")

    for row in ws.iter_rows(min_row=header_row, max_row=header_row + len(detail_tabs), max_col=len(headers)):
        for cell in row:
            cell.border = BORDER
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(detail_tabs)}"


def _write_detail_tab(wb, sname, row):
    ws = wb.create_sheet(title=sname)
    ws.cell(row=1, column=1, value=f"Folder: {row.get('folder', '')}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Status: {row.get('status', '')}").font = Font(size=10, bold=True)
    ws.cell(row=3, column=1, value=f"Prod PDF: {row.get('prod_pdf', '')}")
    ws.cell(row=4, column=1, value=f"Stage PDF: {row.get('stage_pdf', '')}")
    ws.cell(row=5, column=1, value=f"Report PDF:")
    report_path = row.get("report", "")
    if report_path:
        _hyperlink_cell(ws, 5, 2, Path(report_path).name, Path(report_path).as_uri())

    start = 7
    headers = ["Field", "Value"]
    widths = [22, 90]
    _style_header(ws, start, headers, widths)
    fields = [
        ("Folder", row.get("folder", "")),
        ("Status", row.get("status", "")),
        ("Prod PDF", row.get("prod_pdf", "")),
        ("Stage PDF", row.get("stage_pdf", "")),
        ("TOC match", row.get("toc_match", "")),
        ("TOC missing", row.get("toc_missing", "")),
        ("TOC extra", row.get("toc_extra", "")),
        ("Content pass", row.get("content_pass", "")),
        ("Content fail", row.get("content_fail", "")),
        ("Report path", report_path),
    ]
    for i, (label, value) in enumerate(fields, start=start + 1):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=i, column=2, value=value)
        cell.alignment = WRAP_TOP
        cell.border = BORDER
    ws.column_dimensions[get_column_letter(1)].width = 22
    ws.column_dimensions[get_column_letter(2)].width = 90
    ws.freeze_panes = "A8"


if __name__ == "__main__":
    main()
