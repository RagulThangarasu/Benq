#!/usr/bin/env python3
"""Folder-wise STYLE validation -> single Excel workbook.

PROD (expected) = "Final Cleanup .../prod_pdfs/<folder>"
STAGE (actual)  = "Benq/benq_pdfs/<folder>"

For every product folder present in both locations, run style_validation
(validate_style) and write the findings into one .xlsx:
  * Tab 1  "Consolidated" — one row per product with per-category counts +
                            High/Medium/Low totals (links to each tab).
  * Tabs 2..N            — one tab per product with its detailed findings.

Output: content_validation/folderwise_reports/style_validation_folderwise.xlsx
"""
import sys
import re
from pathlib import Path

HERE = Path(__file__).resolve().parent          # .../Benq/content_validation
ROOT = HERE.parent                               # .../Benq
sys.path.insert(0, str(HERE))

from style_validation import validate_style, CATEGORY_ORDER   # noqa: E402

import openpyxl                                               # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROD_BASE = Path(
    "/Users/ragul/Desktop/Final Cleanup files for Pavan to add Structure "
    "and fix the alt attribute value/prod_pdfs"
)
STAGE_BASE = ROOT / "benq_pdfs"
OUT_XLSX = HERE / "folderwise_reports" / "style_validation_folderwise.xlsx"

SEV_FILL = {
    "High":   PatternFill("solid", fgColor="F8CBAD"),
    "Medium": PatternFill("solid", fgColor="FFE699"),
    "Low":    PatternFill("solid", fgColor="BDD7EE"),
}
HDR_FILL = PatternFill("solid", fgColor="34495E")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def _first_pdf(folder: Path):
    pdfs = sorted(p for p in folder.glob("*.pdf") if not p.name.startswith("._"))
    return pdfs[0] if pdfs else None


def _sheet_name(name, used):
    safe = re.sub(r"[\[\]:*?/\\]", "-", name).strip()[:31] or "sheet"
    base, i = safe, 1
    while safe.lower() in used:
        suffix = f"~{i}"
        safe = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(safe.lower())
    return safe


def _style_header(ws, row, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w


def main():
    OUT_XLSX.parent.mkdir(exist_ok=True)
    prod_folders = {p.name for p in PROD_BASE.iterdir() if p.is_dir()}
    stage_folders = {p.name for p in STAGE_BASE.iterdir() if p.is_dir()}
    common = sorted(prod_folders & stage_folders)
    print(f"Common folders: {len(common)}")

    wb = openpyxl.Workbook()
    cons = wb.active
    cons.title = "Consolidated"

    # run validation + create one detail tab per folder ----------------------
    used_names = {"consolidated"}
    rows = []  # (folder, prod, stage, sheet_name, findings|None, error)
    for i, name in enumerate(common, 1):
        prod_pdf = _first_pdf(PROD_BASE / name)
        stage_pdf = _first_pdf(STAGE_BASE / name)
        print(f"[{i}/{len(common)}] {name}")
        if not prod_pdf or not stage_pdf:
            rows.append((name, "-", "-", None, None, "missing PDF"))
            continue
        try:
            findings = validate_style(str(prod_pdf), str(stage_pdf))
        except Exception as e:
            print(f"   ERROR: {e}")
            rows.append((name, prod_pdf.name, stage_pdf.name, None, None, str(e)[:150]))
            continue
        sname = _sheet_name(name, used_names)
        _write_detail_tab(wb, sname, name, prod_pdf, stage_pdf, findings)
        rows.append((name, prod_pdf.name, stage_pdf.name, sname, findings, None))
        sev = {s: sum(1 for f in findings if f["severity"] == s)
               for s in ("High", "Medium", "Low")}
        print(f"   {len(findings)} findings (H{sev['High']} M{sev['Medium']} L{sev['Low']})")

    _write_consolidated(cons, rows)
    wb.save(OUT_XLSX)
    print(f"\nSaved: {OUT_XLSX}")
    print(f"Tabs: 1 consolidated + {sum(1 for r in rows if r[3])} folder tabs")


def _write_consolidated(ws, rows):
    ws.cell(row=1, column=1, value="Style Validation — Consolidated Report "
            "(PROD = expected / STAGE = actual)").font = TITLE_FONT
    headers = (["#", "Product folder", "Total", "High", "Medium", "Low"]
               + CATEGORY_ORDER)
    widths = ([4, 34, 7, 7, 8, 7] + [13] * len(CATEGORY_ORDER))
    _style_header(ws, 3, headers, widths)

    r = 4
    for idx, (name, prod, stage, sname, findings, err) in enumerate(rows, 1):
        ws.cell(row=r, column=1, value=idx)
        # folder name links to its detail tab when available
        fcell = ws.cell(row=r, column=2)
        if sname:
            fcell.value = f'=HYPERLINK("#\'{sname}\'!A1","{name}")'
            fcell.font = Font(color="0563C1", underline="single")
        else:
            fcell.value = f"{name}  ({err})"
            fcell.font = Font(italic=True, color="C00000")

        if findings is None:
            for c in range(3, 7 + len(CATEGORY_ORDER)):
                ws.cell(row=r, column=c, value="—").alignment = CENTER
            r += 1
            continue

        sev = {s: sum(1 for f in findings if f["severity"] == s)
               for s in ("High", "Medium", "Low")}
        ws.cell(row=r, column=3, value=len(findings)).alignment = CENTER
        for c, s in ((4, "High"), (5, "Medium"), (6, "Low")):
            cell = ws.cell(row=r, column=c, value=sev[s])
            cell.alignment = CENTER
            if sev[s]:
                cell.fill = SEV_FILL[s]
        for j, cat in enumerate(CATEGORY_ORDER):
            n = sum(1 for f in findings if f["category"] == cat)
            cell = ws.cell(row=r, column=7 + j, value=n if n else "")
            cell.alignment = CENTER
        r += 1

    # borders across the table
    for row in ws.iter_rows(min_row=3, max_row=r - 1, max_col=6 + len(CATEGORY_ORDER)):
        for cell in row:
            cell.border = BORDER
    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:{get_column_letter(6 + len(CATEGORY_ORDER))}{r - 1}"


def _write_detail_tab(wb, sname, folder, prod_pdf, stage_pdf, findings):
    ws = wb.create_sheet(title=sname)
    ws.cell(row=1, column=1, value=folder).font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"PROD (expected): {prod_pdf.name}").font = Font(size=9, color="666666")
    ws.cell(row=3, column=1, value=f"STAGE (actual):  {stage_pdf.name}").font = Font(size=9, color="666666")
    ws.cell(row=4, column=1,
            value=(f"{len(findings)} findings  —  "
                   + ", ".join(f"{s}:{sum(1 for f in findings if f['severity']==s)}"
                               for s in ('High', 'Medium', 'Low')))
            ).font = Font(size=9, bold=True)
    ws.cell(row=5, column=1, value="")

    headers = ["#", "Category", "Severity", "Topic", "Pages",
               "Expected (PROD)", "Actual (STAGE)", "Issue", "Fix"]
    widths = [4, 18, 9, 26, 10, 30, 30, 40, 40]
    _style_header(ws, 6, headers, widths)

    if not findings:
        ws.cell(row=7, column=2, value="No style issues found.").font = Font(italic=True, color="2E7D32")
        ws.freeze_panes = "A7"
        return

    order = {c: i for i, c in enumerate(CATEGORY_ORDER)}
    sev_rank = {"High": 0, "Medium": 1, "Low": 2}
    fsorted = sorted(findings, key=lambda f: (order.get(f["category"], 99),
                                              sev_rank.get(f["severity"], 9)))
    r = 7
    for idx, f in enumerate(fsorted, 1):
        vals = [idx, f["category"], f["severity"], f.get("topic", ""),
                str(f.get("pages", "")), f.get("expected", ""),
                f.get("actual", ""), f.get("issue", ""), f.get("fix", "")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = WRAP_TOP
            cell.border = BORDER
        sc = ws.cell(row=r, column=3)
        sc.alignment = CENTER
        if f["severity"] in SEV_FILL:
            sc.fill = SEV_FILL[f["severity"]]
        r += 1
    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:I{r - 1}"


if __name__ == "__main__":
    main()
